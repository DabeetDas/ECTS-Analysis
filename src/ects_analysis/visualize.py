from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any


BANK_LABELS = {
    "BOB": "Bank of Baroda",
    "SBI": "State Bank of India",
    "PNB": "Punjab National Bank",
    "BOI": "Bank of India",
    "UBOI": "Union Bank of India",
    "CAN": "Canara Bank",
    "IB": "Indian Bank",
}

BANK_BRANDS = {
    "BOB": {
        "primary": "#C27D38",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
    "SBI": {
        "primary": "#1E6B7B",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
    "PNB": {
        "primary": "#8A6A3F",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
    "BOI": {
        "primary": "#0D5EAF",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
    "UBOI": {
        "primary": "#D52B1E",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
    "CAN": {
        "primary": "#005696",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
    "IB": {
        "primary": "#003A70",
        "secondary": "#0F1E36",
        "accent": "#F5F0E6",
    },
}

EXCEL_BANK_SHEETS = {
    "BOB": "Bank of Baroda",
    "SBI": "SBI",
    "BOI": "Bank of India",
    "UBOI": "Union Bank of India",
    "CAN": "Canara Bank",
    "IB": "Indian Bank",
}

EXCEL_RATIO_FIELDS = {
    "bank z score": ("z_score", "number"),
    "capital to risk weight assets ratio (crar)": ("crar_pct", "percent"),
    "casa ratio": ("casa_pct", "percent"),
    "cet1 ratio": ("cet1_pct", "percent"),
    "cost to income ratio": ("cost_to_income_pct", "percent"),
    "credit deposit ratio": ("credit_deposit_pct", "percent"),
    "efficiency ratio": ("efficiency_ratio_pct", "percent"),
    "equity / assets": ("equity_assets_pct", "percent"),
    "gross non performing assets (gnpa)": ("gnpa_pct", "percent"),
    "liquidity coverage ratio": ("lcr_pct", "percent"),
    "loan to assets ratio": ("loan_to_assets_pct", "percent"),
    "net interest margin": ("nim_pct", "percent"),
    "net non performing assets (nnpa)": ("nnpa_pct", "percent"),
    "operating leverage": ("operating_leverage_pct", "percent"),
    "provision coverage ratio (pcr) including auca": ("pcr_pct", "percent"),
    "return on assets ratio": ("roa_pct", "percent"),
    "return on equity ratio": ("roe_pct", "percent"),
    "roa": ("roa_pct", "percent"),
    "volatility": ("roa_volatility_pct", "percent"),
    "z-score": ("z_score", "number"),
}

EXCEL_HIGH_LEVEL_FIELDS = {
    ("income", "interest"): "interest_income_cr",
    ("income", "other"): "other_income_cr",
    ("income", "total"): "total_income_cr",
    ("expenditure", "interest expended"): "interest_expended_cr",
    ("expenditure", "operating expenses"): "operating_expenses_cr",
    ("expenditure", "provisions and contingencies"): "provisions_contingencies_cr",
    ("expenditure", "total"): "total_expenditure_cr",
    ("profit", "net profit"): "profit_after_tax_cr",
}


def load_corpus_analysis(path: Path) -> dict[str, Any]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Corpus analysis JSON must be an object")

    required_keys = ["trend_analysis", "competitor_analysis", "observations"]
    missing = [key for key in required_keys if key not in payload]
    if missing:
        raise ValueError(f"Corpus analysis JSON is missing: {', '.join(missing)}")
    if not isinstance(payload.get("competitor_analysis"), dict):
        raise ValueError("Corpus analysis `competitor_analysis` must be an object")
    return payload


def load_financial_analysis(path: Path) -> dict[str, list[dict[str, str]]]:
    if not path.exists():
        raise FileNotFoundError(f"Financial file not found: {path}")

    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        return load_financial_analysis_from_excel(path)

    with path.open(newline="", encoding="utf-8") as csv_file:
        rows = list(csv.DictReader(csv_file))

    if not rows:
        return {}

    required_fields = {"bank", "fiscal_year"}
    missing_fields = required_fields - set(rows[0])
    if missing_fields:
        raise ValueError(f"Financial CSV is missing: {', '.join(sorted(missing_fields))}")

    by_bank: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        bank = str(row.get("bank", "")).strip()
        if not bank:
            continue
        normalized = {str(key): str(value or "").strip() for key, value in row.items() if key is not None}
        by_bank[bank].append(normalized)

    return {
        bank: sorted(bank_rows, key=lambda item: fiscal_year_sort_key(item.get("fiscal_year", "")))
        for bank, bank_rows in by_bank.items()
    }


def load_financial_analysis_from_excel(path: Path) -> dict[str, list[dict[str, str]]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("Reading Excel financials requires the `openpyxl` package") from exc

    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    high_level = load_excel_high_level_comparison(workbook)
    by_bank: dict[str, dict[str, dict[str, str]]] = defaultdict(dict)

    for bank, sheet_name in EXCEL_BANK_SHEETS.items():
        if sheet_name not in workbook.sheetnames:
            continue
        sheet_rows = list(workbook[sheet_name].iter_rows(values_only=True))
        header_index, fiscal_year_columns = find_excel_year_columns(sheet_rows)
        if header_index is None:
            continue

        for row in sheet_rows[header_index + 1 :]:
            label_index = min(fiscal_year_columns) - 1
            if label_index < 0 or label_index >= len(row):
                continue
            label = normalize_excel_label(row[label_index])
            field = EXCEL_RATIO_FIELDS.get(label)
            if not field:
                continue
            field_name, value_type = field
            for column_index, fiscal_year in fiscal_year_columns.items():
                if column_index >= len(row):
                    continue
                value = row[column_index]
                if not is_number(value):
                    continue
                bank_row = by_bank[bank].setdefault(
                    fiscal_year,
                    {"bank": bank, "fiscal_year": fiscal_year},
                )
                bank_row[field_name] = format_excel_value(float(value), value_type)

        for fiscal_year, fields in high_level.get(bank, {}).items():
            bank_row = by_bank[bank].setdefault(
                fiscal_year,
                {"bank": bank, "fiscal_year": fiscal_year},
            )
            bank_row.update(fields)

    return {
        bank: sorted(bank_rows.values(), key=lambda item: fiscal_year_sort_key(item.get("fiscal_year", "")))
        for bank, bank_rows in by_bank.items()
    }


def load_excel_high_level_comparison(workbook: Any) -> dict[str, dict[str, dict[str, str]]]:
    if "High level comparison" not in workbook.sheetnames:
        return {}

    rows = list(workbook["High level comparison"].iter_rows(values_only=True))
    bank_columns: dict[int, str] = {}
    for row in rows[:5]:
        for column_index, value in enumerate(row):
            text = normalize_excel_label(value)
            for bank, sheet_name in EXCEL_BANK_SHEETS.items():
                if text == normalize_excel_label(sheet_name):
                    bank_columns[column_index] = bank
        if bank_columns:
            break

    by_bank: dict[str, dict[str, dict[str, str]]] = defaultdict(lambda: defaultdict(dict))
    current_category = ""
    for row in rows:
        if len(row) > 1 and row[1]:
            current_category = normalize_excel_label(row[1])
        if len(row) <= 2 or not current_category:
            continue
        line_item = normalize_excel_label(row[2])
        field_name = EXCEL_HIGH_LEVEL_FIELDS.get((current_category, line_item))
        if not field_name:
            continue
        for column_index, bank in bank_columns.items():
            if column_index >= len(row) or not is_number(row[column_index]):
                continue
            by_bank[bank]["FY25"][field_name] = format_crore(float(row[column_index]) / 10_000_000)

    return {bank: dict(years) for bank, years in by_bank.items()}


def find_excel_year_columns(
    rows: list[tuple[Any, ...]],
) -> tuple[int | None, dict[int, str]]:
    for row_index, row in enumerate(rows):
        fiscal_year_columns: dict[int, str] = {}
        for column_index, value in enumerate(row):
            fiscal_year = fiscal_year_from_excel_header(value)
            if fiscal_year:
                fiscal_year_columns[column_index] = fiscal_year
        if fiscal_year_columns:
            return row_index, fiscal_year_columns
    return None, {}


def fiscal_year_from_excel_header(value: Any) -> str | None:
    text = str(value or "")
    if "as at" not in text.casefold():
        return None
    match = re.search(r"\b20(\d{2})\b", text)
    return f"FY{match.group(1)}" if match else None


def normalize_excel_label(value: Any) -> str:
    return " ".join(str(value or "").casefold().strip().split())


def is_number(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def format_excel_value(value: float, value_type: str) -> str:
    if value_type == "percent":
        return format_crore(value if abs(value) > 2 else value * 100)
    return format_crore(value)


def format_crore(value: float) -> str:
    rounded = round(value, 2)
    return str(int(rounded)) if rounded.is_integer() else f"{rounded:.2f}"


def load_topic_hierarchy(path: Path) -> dict[str, list[str]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(payload, dict):
        raise ValueError("Topic hierarchy JSON must be an object")

    hierarchy: dict[str, list[str]] = {}
    for topic, subtopics in payload.items():
        if not isinstance(topic, str) or not isinstance(subtopics, list):
            raise ValueError("Topic hierarchy must map topic names to subtopic lists")
        if not all(isinstance(subtopic, str) for subtopic in subtopics):
            raise ValueError("Topic hierarchy subtopic lists must contain only strings")
        hierarchy[topic] = subtopics
    return hierarchy


def export_dashboard_payload(
    analysis: dict[str, Any],
    *,
    financial_analysis: dict[str, list[dict[str, str]]] | None = None,
    topic_hierarchy: dict[str, list[str]] | None = None,
    profile_bank: str = "BOB",
) -> dict[str, Any]:
    companies = sorted(
        {
            *[
                str(document.get("company"))
                for document in _list(analysis.get("documents"))
                if isinstance(document, dict) and document.get("company")
            ],
            *[
                str(observation.get("company"))
                for observation in _list(analysis.get("observations"))
                if isinstance(observation, dict) and observation.get("company")
            ],
            *[str(company) for company in (financial_analysis or {}) if company],
        },
        key=bank_sort_key,
    )

    return {
        "metadata": {
            "title": "PSB Earnings Call Intelligence",
            "profile_bank": profile_bank,
            "generated_from": "ects_analysis.visualize",
        },
        "banks": {
            company: {
                "code": company,
                "label": BANK_LABELS.get(company, company),
                "brand": BANK_BRANDS.get(
                    company,
                    {"primary": "#5C6773", "secondary": "#0F1E36", "accent": "#F5F0E6"},
                ),
            }
            for company in companies
        },
        "topic_hierarchy": topic_hierarchy or {},
        "financial_analysis": financial_analysis or {},
        "analysis": analysis,
    }


def write_dashboard_payload(
    *,
    input_path: Path,
    output_path: Path,
    financials_path: Path | None,
    topic_hierarchy_path: Path | None,
    profile_bank: str,
) -> None:
    analysis = load_corpus_analysis(input_path)
    analysis = preserve_existing_analysis_when_input_is_empty(analysis, output_path)
    financial_analysis = load_financial_analysis(financials_path) if financials_path else {}
    topic_hierarchy = load_topic_hierarchy(topic_hierarchy_path) if topic_hierarchy_path else {}
    payload = export_dashboard_payload(
        analysis,
        financial_analysis=financial_analysis,
        topic_hierarchy=topic_hierarchy,
        profile_bank=profile_bank,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")


def preserve_existing_analysis_when_input_is_empty(
    analysis: dict[str, Any],
    output_path: Path,
) -> dict[str, Any]:
    if observation_count(analysis) > 0 or not output_path.exists():
        return analysis

    try:
        existing_payload = json.loads(output_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return analysis

    if not isinstance(existing_payload, dict):
        return analysis
    existing_analysis = existing_payload.get("analysis")
    if isinstance(existing_analysis, dict) and observation_count(existing_analysis) > 0:
        return existing_analysis
    return analysis


def observation_count(analysis: dict[str, Any]) -> int:
    return len(_list(analysis.get("observations")))


def fiscal_year_sort_key(value: str) -> tuple[int, str]:
    digits = "".join(character for character in str(value) if character.isdigit())
    return (int(digits) if digits else 0, str(value))


def bank_sort_key(value: str) -> tuple[str, str]:
    code = str(value)
    return (BANK_LABELS.get(code, code).casefold(), code.casefold())


def _list(value: Any) -> list[Any]:
    return value if isinstance(value, list) else []


def main() -> None:
    parser = argparse.ArgumentParser(description="Export dashboard data for the Next.js frontend.")
    parser.add_argument("--input", default="outputs/psb_corpus_analysis.json", help="Corpus analysis JSON")
    parser.add_argument(
        "--output",
        default="frontend/public/data/dashboard-data.json",
        help="Output JSON consumed by the frontend",
    )
    parser.add_argument(
        "--financials",
        default="data_new.xlsx",
        help="Financial CSV or Excel workbook to overlay in the frontend",
    )
    parser.add_argument(
        "--topic-hierarchy",
        default="data/psb_seed_topics.json",
        help="Topic/subtopic hierarchy JSON",
    )
    parser.add_argument("--profile-bank", default="BOB", help="Featured bank code")
    parser.add_argument(
        "--view",
        choices=["corpus"],
        default="corpus",
        help="Kept for CLI compatibility; the React frontend renders the view.",
    )
    args = parser.parse_args()

    write_dashboard_payload(
        input_path=Path(args.input),
        output_path=Path(args.output),
        financials_path=Path(args.financials) if args.financials else None,
        topic_hierarchy_path=Path(args.topic_hierarchy) if args.topic_hierarchy else None,
        profile_bank=args.profile_bank,
    )
    print(f"Wrote frontend dashboard data to {args.output}")


if __name__ == "__main__":
    main()
